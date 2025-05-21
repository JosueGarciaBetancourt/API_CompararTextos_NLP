
import os
import json
import requests
import time
import numpy as np
import pandas as pd
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    mean_squared_error, mean_absolute_error, r2_score,
    confusion_matrix, classification_report
)
import matplotlib.pyplot as plt
import seaborn as sns
from collections import defaultdict
from tqdm import tqdm
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.chart import BarChart, Reference, LineChart, ScatterChart
from openpyxl.chart.series import Series
from openpyxl.drawing.image import Image
from io import BytesIO

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("semantic_precision_evaluation.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("semantic_evaluator")

# Configuración
API_URL = "http://localhost:5000/comparar_cursos"
TEST_JSONS_DIR = "jsons_prueba_UC_UNMSM"
RESULTS_DIR = "resultados_evaluacion"
GOLDEN_STANDARD_FILE = "golden_standard/golden_standard_convalidaciones.json"
THRESHOLD_SIMILITUD_PONDERADA = 0.70
THRESHOLD_HIGH_SIMILARITY = 0.85
THRESHOLD_MEDIUM_SIMILARITY = 0.75
NUM_THREADS = 4

# Pesos para calcular la similitud ponderada
PESOS_SIMILITUD = {
    "similitud_sumilla": 0.25,
    "similitud_aprendizajes": 0.10,
    "similitud_unidades": 0.60,
    "similitud_bibliografia": 0.05
}

# Crear directorio de resultados si no existe
if not os.path.exists(RESULTS_DIR):
    os.makedirs(RESULTS_DIR)

class ExcelReportGenerator:
    """Clase para generar reportes en Excel con formato profesional."""

    def __init__(self, filename):
        self.filename = filename
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.styles = self._define_styles()

    def _define_styles(self):
        """Define estilos para el reporte."""
        return {
            'header': Font(name='Calibri', bold=True, size=12, color='FFFFFF'),
            'header_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
            'title': Font(name='Calibri', bold=True, size=14),
            'subtitle': Font(name='Calibri', italic=True, size=12),
            'highlight': Font(bold=True, color='FF0000'),
            'good': Font(color='00B050'),
            'bad': Font(color='FF0000'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'wrap': Alignment(wrap_text=True),
            'green_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'red_fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'yellow_fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        }

    def add_sheet(self, title, df, notes=None, chart_data=None, conditional_cols=None):
        """Añade una hoja con datos y formato."""
        try:
            # Verificar si el DataFrame está vacío
            if df.empty or len(df.columns) == 0:
                logger.warning(f"DataFrame vacío para la hoja '{title}'")
                df = pd.DataFrame({"Mensaje": ["No hay datos disponibles"]})

            ws = self.wb.create_sheet(title=title)

            # Añadir título solo si hay columnas
            if len(df.columns) > 0:
                ws.append([title])
                # Asegurarse de que end_column no sea menor que start_column
                end_col = max(1, len(df.columns))
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
                ws['A1'].font = self.styles['title']
                ws['A1'].alignment = self.styles['center']

            # Añadir datos
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Formato encabezados (si hay datos)
            if len(df.columns) > 0 and ws.max_row >= 2:
                for cell in ws[2]:
                    cell.font = self.styles['header']
                    cell.fill = self.styles['header_fill']
                    cell.alignment = self.styles['center']
                    cell.border = self.styles['border']

            # Formato celdas de datos
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = self.styles['border']
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

            # Formato condicional
            if conditional_cols:
                for col, rules in conditional_cols.items():
                    try:
                        col_idx = df.columns.get_loc(col) + 1
                        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                for rule in rules:
                                    if rule['type'] == 'greater_equal' and cell.value >= rule['value']:
                                        cell.fill = self.styles['green_fill']
                                    elif rule['type'] == 'less' and cell.value < rule['value']:
                                        cell.fill = self.styles['red_fill']
                                    elif rule['type'] == 'range' and rule['min'] <= cell.value < rule['max']:
                                        cell.fill = self.styles['yellow_fill']
                    except KeyError:
                        logger.warning(f"Columna '{col}' no encontrada para formato condicional")

            # Autoajustar columnas
            for col in ws.columns:
                # Saltar celdas fusionadas
                if col[0].__class__.__name__ == 'MergedCell':
                    continue

                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Añadir notas
            if notes:
                ws.append([])
                for note in notes:
                    ws.append([note])

            # Añadir gráfico si hay datos suficientes
            if chart_data and len(df) > 1:
                self._add_chart(ws, df, chart_data)

        except Exception as e:
            logger.error(f"Error al añadir hoja '{title}': {str(e)}")
            raise

    def _add_chart(self, ws, df, chart_data):
        """Añade un gráfico a la hoja."""
        chart_type = chart_data.get('type', 'bar')

        if chart_type == 'bar':
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = chart_data.get('title', '')
            chart.y_axis.title = chart_data.get('y_title', '')
            chart.x_axis.title = chart_data.get('x_title', '')

            data = Reference(ws, min_col=2, min_row=2, max_row=len(df)+2, max_col=len(df.columns))
            cats = Reference(ws, min_col=1, min_row=3, max_row=len(df)+2)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, chart_data.get('pos', 'A{}'.format(len(df) + 5)))

        elif chart_type == 'line':
            chart = LineChart()
            chart.title = chart_data.get('title', '')
            chart.y_axis.title = chart_data.get('y_title', '')
            chart.x_axis.title = chart_data.get('x_title', '')

            data = Reference(ws, min_col=2, min_row=2, max_row=len(df)+2, max_col=len(df.columns))
            cats = Reference(ws, min_col=1, min_row=3, max_row=len(df)+2)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, chart_data.get('pos', 'A{}'.format(len(df) + 5)))

        elif chart_type == 'scatter':
            chart = ScatterChart()
            chart.title = chart_data.get('title', '')
            chart.y_axis.title = chart_data.get('y_title', '')
            chart.x_axis.title = chart_data.get('x_title', '')

            x_values = Reference(ws, min_col=chart_data['x_col'], min_row=3, max_row=len(df)+2)
            y_values = Reference(ws, min_col=chart_data['y_col'], min_row=3, max_row=len(df)+2)

            series = Series(y_values, x_values, title_from_data=True)
            chart.series.append(series)

            # Línea de referencia y=x
            ref_line = Reference(ws, min_col=1, min_row=3, max_row=len(df)+2)
            chart.series.append(Series(ref_line, ref_line, title="Referencia"))

            ws.add_chart(chart, chart_data.get('pos', 'A{}'.format(len(df) + 5)))

    def add_image(self, sheet_name, image_data, pos):
        """Añade una imagen a la hoja especificada."""
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            img = Image(BytesIO(image_data))
            ws.add_image(img, pos)

    def save(self):
        """Guarda el archivo Excel."""
        try:
            self.wb.save(self.filename)
        except Exception as e:
            logger.error(f"Error al guardar el archivo Excel: {str(e)}")
            raise


class SemanticPrecisionEvaluator:
    """Evaluador de precisión semántica para API de comparación de cursos."""

    def __init__(self):
        self.resultados = []
        self.metricas_globales = {}
        self.golden_standard = self.cargar_golden_standard()
        self.report_generator = ExcelReportGenerator(
            os.path.join(RESULTS_DIR, "reporte_evaluacion.xlsx")
        )
        # Precisión para comparaciones con el umbral
        self.PRECISION = 2  # Número de decimales para redondear antes de comparar

    def es_similar(self, valor):
        """Determina si un valor de similitud supera el umbral, con redondeo para evitar errores de punto flotante."""
        return round(valor, self.PRECISION) >= round(THRESHOLD_SIMILITUD_PONDERADA, self.PRECISION)

    def cargar_golden_standard(self):
        """Carga el estándar dorado para comparación con múltiples claves de búsqueda."""
        try:
            with open(GOLDEN_STANDARD_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                golden_dict = {}
                for comp in data.get("comparaciones", []):
                    # Crear múltiples claves para flexibilidad en la búsqueda
                    id_origen = comp.get('idCursoOrigen')
                    id_destino = comp.get('idCursoDestino')
                    nombre_origen = comp.get('nombreCursoOrigen', '').lower().strip()
                    nombre_destino = comp.get('nombreCursoDestino', '').lower().strip()

                    # Clave con IDs
                    if id_origen is not None and id_destino is not None:
                        golden_dict[f"{id_origen}_{id_destino}"] = comp.get("resultado_resumido", {})

                    # Clave con nombres
                    if nombre_origen and nombre_destino:
                        golden_dict[f"{nombre_origen}_{nombre_destino}"] = comp.get("resultado_resumido", {})

                    # Clave solo con ID origen y nombre destino (para casos mixtos)
                    if id_origen is not None and nombre_destino:
                        golden_dict[f"{id_origen}_{nombre_destino}"] = comp.get("resultado_resumido", {})

                    # Clave solo con nombre origen e ID destino
                    if nombre_origen and id_destino is not None:
                        golden_dict[f"{nombre_origen}_{id_destino}"] = comp.get("resultado_resumido", {})

                logger.info(f"Golden standard cargado con {len(golden_dict)} claves de búsqueda")
                return golden_dict
        except Exception as e:
            logger.error(f"Error al cargar golden standard: {str(e)}")
            return {}

    def enviar_json_a_api(self, json_data):
        """Envía un JSON a la API y retorna la respuesta."""
        try:
            start_time = time.time()
            response = requests.post(API_URL, json=json_data, timeout=600)
            response_time = time.time() - start_time

            if response.status_code != 200:
                logger.error(f"API respondió con código {response.status_code}: {response.text}")
                return None, response_time

            return response.json(), response_time
        except requests.exceptions.Timeout:
            logger.error("Timeout al conectar con la API (10 minutos)")
            return None, 300
        except Exception as e:
            logger.error(f"Error al enviar solicitud a la API: {str(e)}")
            return None, 0

    def cargar_json_prueba(self, json_file):
        """Carga un archivo JSON de prueba."""
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error al cargar JSON {json_file}: {str(e)}")
            return None

    def calcular_similitud_ponderada(self, similitudes):
        """Calcula la similitud ponderada."""
        try:
            total_pesos = sum(peso for componente, peso in PESOS_SIMILITUD.items() if componente in similitudes)
            if total_pesos == 0:
                return 0.0

            return sum(
                similitudes[componente] * peso
                for componente, peso in PESOS_SIMILITUD.items()
                if componente in similitudes
            ) / total_pesos
        except Exception as e:
            logger.error(f"Error al calcular similitud ponderada: {str(e)}")
            return 0.0

    def evaluar_precision_similitud(self, predicciones, referencias=None):
        """Evalúa la precisión de las predicciones de similitud."""
        if not predicciones:
            return {
                "mean": 0, "median": 0, "std": 0, "min": 0, "max": 0,
                "count_high_similarity": 0, "count_medium_similarity": 0, "count_low_similarity": 0,
                "percent_high_similarity": 0, "percent_medium_similarity": 0, "percent_low_similarity": 0
            }

        try:
            predicciones = np.array(predicciones)
            resultados = {
                "mean": float(np.mean(predicciones)),
                "median": float(np.median(predicciones)),
                "std": float(np.std(predicciones)),
                "min": float(np.min(predicciones)),
                "max": float(np.max(predicciones)),
                "count_high_similarity": int(np.sum(predicciones >= THRESHOLD_HIGH_SIMILARITY)),
                "count_medium_similarity": int(np.sum((predicciones >= THRESHOLD_MEDIUM_SIMILARITY) &
                                                      (predicciones < THRESHOLD_HIGH_SIMILARITY))),
                "count_low_similarity": int(np.sum(predicciones < THRESHOLD_MEDIUM_SIMILARITY)),
            }

            total = len(predicciones)
            for k in ['high', 'medium', 'low']:
                resultados[f"percent_{k}_similarity"] = float(
                    resultados[f"count_{k}_similarity"] / total * 100) if total > 0 else 0

            if referencias is not None and len(referencias) == len(predicciones):
                referencias = np.array(referencias)
                resultados.update({
                    "mae": float(mean_absolute_error(referencias, predicciones)),
                    "mse": float(mean_squared_error(referencias, predicciones)),
                    "rmse": float(np.sqrt(mean_squared_error(referencias, predicciones))),
                    "r2": float(r2_score(referencias, predicciones)),
                    "accuracy": float(accuracy_score(
                        (referencias >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        (predicciones >= THRESHOLD_SIMILITUD_PONDERADA).astype(int)
                    )),
                    "precision": float(precision_score(
                        (referencias >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        (predicciones >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        zero_division=0
                    )),
                    "recall": float(recall_score(
                        (referencias >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        (predicciones >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        zero_division=0
                    )),
                    "f1": float(f1_score(
                        (referencias >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        (predicciones >= THRESHOLD_SIMILITUD_PONDERADA).astype(int),
                        zero_division=0
                    ))
                })

            return resultados
        except Exception as e:
            logger.error(f"Error en evaluar_precision_similitud: {str(e)}")
            return {
                "mean": 0, "median": 0, "std": 0, "min": 0, "max": 0,
                "count_high_similarity": 0, "count_medium_similarity": 0, "count_low_similarity": 0,
                "percent_high_similarity": 0, "percent_medium_similarity": 0, "percent_low_similarity": 0
            }

    def procesar_json_prueba(self, json_file):
        """Procesa un archivo JSON de prueba y evalúa el resultado."""
        nombre_archivo = os.path.basename(json_file)
        json_data = self.cargar_json_prueba(json_file)

        if json_data is None:
            return None

        resultado, tiempo_respuesta = self.enviar_json_a_api(json_data)

        if resultado is None:
            return None

        evaluacion = {
            "nombre_archivo": nombre_archivo,
            "tiempo_respuesta": tiempo_respuesta,
            "num_comparaciones": len(resultado.get("comparaciones", [])),
            "metricas_similitud": {},
            "metricas_emparejamiento": {},
            "metricas_coherencia": {},
            "comparaciones_evaluadas": []
        }

        similitudes = defaultdict(list)
        referencias_ponderadas = []
        temas_comunes_list = []

        for comp in resultado.get("comparaciones", []):
            id_origen = comp.get("idCursoOrigen")
            id_destino = comp.get("idCursoDestino")
            nombre_origen = comp.get("nombreCursoOrigen", "").lower().strip()
            nombre_destino = comp.get("nombreCursoDestino", "").lower().strip()

            if "resultado_resumido" in comp:
                resumido = comp["resultado_resumido"]

                # Calcular similitud ponderada
                sim_ponderada = self.calcular_similitud_ponderada({
                    "similitud_sumilla": resumido.get("similitud_sumilla", 0),
                    "similitud_aprendizajes": resumido.get("similitud_aprendizajes", 0),
                    "similitud_unidades": resumido.get("similitud_unidades", 0),
                    "similitud_bibliografia": resumido.get("similitud_bibliografia", 0)
                })

                # Almacenar similitudes por componente
                for componente in PESOS_SIMILITUD:
                    similitudes[componente].append(resumido.get(componente, 0))
                similitudes["ponderada"].append(sim_ponderada)

                # Comparar con golden standard si existe
                comp_eval = {
                    "idCursoOrigen": id_origen,
                    "idCursoDestino": id_destino,
                    "nombreCursoOrigen": nombre_origen,
                    "nombreCursoDestino": nombre_destino,
                    "similitud_sumilla": resumido.get("similitud_sumilla", 0),
                    "similitud_aprendizajes": resumido.get("similitud_aprendizajes", 0),
                    "similitud_unidades": resumido.get("similitud_unidades", 0),
                    "similitud_bibliografia": resumido.get("similitud_bibliografia", 0),
                    "similitud_ponderada": sim_ponderada,
                    "es_similar": self.es_similar(sim_ponderada)
                }

                """print("SIM_PONDERADA: ", comp_eval['similitud_ponderada'])
                print("ES SIMILAR: ", comp_eval['es_similar'])"""

                # Búsqueda flexible en golden standard
                posibles_claves = [
                    f"{id_origen}_{id_destino}",
                    f"{nombre_origen}_{nombre_destino}",
                    f"{id_origen}_{nombre_destino}",
                    f"{nombre_origen}_{id_destino}"
                ]

                ref_encontrada = None
                for clave in posibles_claves:
                    if clave in self.golden_standard:
                        ref_encontrada = self.golden_standard[clave]
                        logger.info(f"Encontrada referencia en golden standard con clave: {clave}")
                        break

                if ref_encontrada:
                    ref_ponderada = self.calcular_similitud_ponderada({
                        "similitud_sumilla": ref_encontrada.get("similitud_sumilla", 0),
                        "similitud_aprendizajes": ref_encontrada.get("similitud_aprendizajes", 0),
                        "similitud_unidades": ref_encontrada.get("similitud_unidades", 0),
                        "similitud_bibliografia": ref_encontrada.get("similitud_bibliografia", 0)
                    })
                    referencias_ponderadas.append(ref_ponderada)
                    comp_eval.update({
                        "referencia_ponderada": ref_ponderada,
                        "diferencia": abs(sim_ponderada - ref_ponderada),
                        "acierto": self.es_similar(sim_ponderada) == self.es_similar(ref_ponderada)
                    })
                else:
                    logger.warning(
                        f"No se encontró referencia en golden standard para: {id_origen}-{id_destino} o {nombre_origen}-{nombre_destino}")

                evaluacion["comparaciones_evaluadas"].append(comp_eval)

            # Extraer temas comunes para análisis
            if "analisis_unidades" in comp:
                for unidad in comp["analisis_unidades"]:
                    if "temas_comunes" in unidad:
                        temas_comunes_list.append(unidad["temas_comunes"])

        # Calcular métricas generales
        for componente in similitudes:
            evaluacion["metricas_similitud"][componente] = self.evaluar_precision_similitud(
                similitudes[componente],
                referencias_ponderadas if componente == "ponderada" and referencias_ponderadas else None
            )

        evaluacion["metricas_emparejamiento"] = {
            "mean_porcentaje_emparejamiento": np.mean([
                comp["resultado_resumido"].get("porcentaje_emparejamiento_unidades", 0)
                for comp in resultado.get("comparaciones", [])
                if "resultado_resumido" in comp
            ]) if any("resultado_resumido" in comp for comp in resultado.get("comparaciones", [])) else 0
        }

        evaluacion["metricas_coherencia"] = {
            "coherencia_interna": np.mean([
                1 - np.std([
                    resumido.get("similitud_sumilla", 0),
                    resumido.get("similitud_aprendizajes", 0),
                    resumido.get("similitud_unidades", 0),
                    resumido.get("similitud_bibliografia", 0)
                ])
                for comp in resultado.get("comparaciones", [])
                if "resultado_resumido" in comp
            ]) if any("resultado_resumido" in comp for comp in resultado.get("comparaciones", [])) else 0
        }

        return evaluacion

    def ejecutar_evaluacion(self):
        """Ejecuta la evaluación completa sobre todos los archivos JSON de prueba."""
        logger.info("Iniciando evaluación de precisión semántica...")

        json_files = [
            os.path.join(TEST_JSONS_DIR, f)
            for f in os.listdir(TEST_JSONS_DIR)
            if f.endswith('.json')
        ]

        if not json_files:
            logger.error(f"No se encontraron archivos JSON en {TEST_JSONS_DIR}")
            return

        logger.info(f"Procesando {len(json_files)} archivos JSON...")

        tiempos_respuesta = []
        with ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
            futures = {
                executor.submit(self.procesar_json_prueba, json_file): json_file
                for json_file in json_files
            }

            for future in tqdm(as_completed(futures), total=len(futures)):
                try:
                    resultado = future.result()
                    if resultado:
                        self.resultados.append(resultado)
                        tiempos_respuesta.append(resultado["tiempo_respuesta"])
                except Exception as e:
                    logger.error(f"Error procesando {futures[future]}: {str(e)}")

        # Calcular métricas globales
        self.calcular_metricas_globales(tiempos_respuesta)

        # Generar reporte Excel
        self.generar_reporte_excel()

        logger.info("Evaluación completada.")

    def calcular_metricas_globales(self, tiempos_respuesta):
        """Calcula métricas globales a partir de todos los resultados."""
        self.metricas_globales["tiempo_respuesta"] = {
            "mean_tiempo": float(np.mean(tiempos_respuesta)) if tiempos_respuesta else 0,
            "median_tiempo": float(np.median(tiempos_respuesta)) if tiempos_respuesta else 0,
            "min_tiempo": float(np.min(tiempos_respuesta)) if tiempos_respuesta else 0,
            "max_tiempo": float(np.max(tiempos_respuesta)) if tiempos_respuesta else 0,
            "std_tiempo": float(np.std(tiempos_respuesta)) if tiempos_respuesta else 0
        }

        # Recopilar similitudes ponderadas y referencias
        similitudes_ponderadas = []
        referencias_ponderadas = []
        aciertos = []

        for resultado in self.resultados:
            for comp in resultado.get("comparaciones_evaluadas", []):
                similitudes_ponderadas.append(comp.get("similitud_ponderada", 0))
                if "referencia_ponderada" in comp:
                    referencias_ponderadas.append(comp["referencia_ponderada"])
                    aciertos.append(comp["acierto"])

        # Calcular métricas globales para similitud ponderada
        self.metricas_globales["similitud_ponderada"] = self.evaluar_precision_similitud(
            similitudes_ponderadas,
            referencias_ponderadas if referencias_ponderadas else None
        )

        # Tasa de acierto global si hay golden standard
        if aciertos:
            self.metricas_globales["tasa_acierto"] = float(np.mean(aciertos))

        # Coherencia interna global
        coherencias = [
            r["metricas_coherencia"]["coherencia_interna"]
            for r in self.resultados
            if "metricas_coherencia" in r
        ]
        self.metricas_globales["coherencia_interna"] = float(np.mean(coherencias)) if coherencias else 0

        # Porcentaje de emparejamiento global
        emparejamientos = [
            r["metricas_emparejamiento"]["mean_porcentaje_emparejamiento"]
            for r in self.resultados
            if "metricas_emparejamiento" in r
        ]
        self.metricas_globales["porcentaje_emparejamiento"] = float(np.mean(emparejamientos)) if emparejamientos else 0

    def generar_reporte_excel(self):
        """Genera un reporte completo en Excel con los resultados."""
        logger.info("Generando reporte en Excel...")

        # 1. Resumen Ejecutivo
        self._generar_resumen_ejecutivo()

        # 2. Métricas Globales
        self._generar_metricas_globales()

        # 3. Comparación con Golden Standard (si existe)
        if self.golden_standard:
            self._generar_comparacion_modelo_vs_golden()
            self._generar_matriz_confusion()
            self._generar_analisis_errores_detallado()

        # 4. Detalle por Archivo
        self._generar_detalle_por_archivo()

        # 5. Comparaciones Individuales
        self._generar_comparaciones_individuales()

        try:
            self.report_generator.save()
            logger.info(f"Reporte guardado en: {os.path.join(RESULTS_DIR, 'reporte_evaluacion.xlsx')}")
        except Exception as e:
            logger.error(f"Error al guardar el reporte: {str(e)}")
            raise

    def _generar_resumen_ejecutivo(self):
        """Genera la hoja de resumen ejecutivo."""
        data = {
            "Total Archivos Evaluados": [len(self.resultados)],
            "Total Comparaciones": [sum(r['num_comparaciones'] for r in self.resultados)],
            "Tasa Acierto Global": [f"{self.metricas_globales.get('tasa_acierto', 0)*100:.2f}%"],
            "MAE (Error Absoluto Medio)": [f"{self.metricas_globales['similitud_ponderada'].get('mae', 0):.2f}"],
            "RMSE (Error Cuadrático Medio)": [f"{self.metricas_globales['similitud_ponderada'].get('rmse', 0):.2f}"],
            "R² (Coeficiente Determinación)": [f"{self.metricas_globales['similitud_ponderada'].get('r2', 0):.2f}"],
            "Tiempo Respuesta Promedio (s)": [f"{self.metricas_globales['tiempo_respuesta'].get('mean_tiempo', 0):.2f}"]
        }

        df = pd.DataFrame(data).T.reset_index()
        df.columns = ['Métrica', 'Valor']

        self.report_generator.add_sheet(
            title="Resumen Ejecutivo",
            df=df,
            notes=[
                "Resumen de las métricas clave de evaluación",
                f"Umbral de similitud: {THRESHOLD_SIMILITUD_PONDERADA}",
                f"Fecha generación: {time.strftime('%Y-%m-%d %H:%M:%S')}"
            ],
            chart_data={
                'type': 'bar',
                'title': 'Métricas Clave',
                'y_title': 'Valor',
                'pos': 'D2'
            } if len(df) > 1 else None
        )

    def _generar_metricas_globales(self):
        """Genera hojas con métricas globales detalladas."""
        # Datos de similitud ponderada
        sim_pond = self.metricas_globales['similitud_ponderada']
        data_similitud = {
            "Métrica": ["Media", "Mediana", "Desviación Estándar", "Mínimo", "Máximo",
                       "MAE", "MSE", "RMSE", "R²", "Accuracy", "Precision", "Recall", "F1"],
            "Valor": [
                sim_pond.get('mean', 0), sim_pond.get('median', 0), sim_pond.get('std', 0),
                sim_pond.get('min', 0), sim_pond.get('max', 0), sim_pond.get('mae', 0),
                sim_pond.get('mse', 0), sim_pond.get('rmse', 0), sim_pond.get('r2', 0),
                sim_pond.get('accuracy', 0), sim_pond.get('precision', 0),
                sim_pond.get('recall', 0), sim_pond.get('f1', 0)
            ]
        }
        df_similitud = pd.DataFrame(data_similitud)

        # Datos de distribución de similitud
        data_distribucion = {
            "Categoría": ["Alta Similitud", "Media Similitud", "Baja Similitud"],
            "Conteo": [
                sim_pond.get('count_high_similarity', 0),
                sim_pond.get('count_medium_similarity', 0),
                sim_pond.get('count_low_similarity', 0)
            ],
            "Porcentaje": [
                sim_pond.get('percent_high_similarity', 0),
                sim_pond.get('percent_medium_similarity', 0),
                sim_pond.get('percent_low_similarity', 0)
            ]
        }
        df_distribucion = pd.DataFrame(data_distribucion)

        # Añadir hojas al reporte
        self.report_generator.add_sheet(
            title="Métricas Similitud",
            df=df_similitud,
            chart_data={'type': 'bar', 'title': 'Métricas de Similitud', 'y_title': 'Valor', 'pos': 'D2'}
        )

        self.report_generator.add_sheet(
            title="Distribución Similitud",
            df=df_distribucion,
            chart_data={'type': 'bar', 'title': 'Distribución de Similitudes', 'y_title': 'Porcentaje', 'pos': 'D2'}
        )

    def _generar_comparacion_modelo_vs_golden(self):
        """Genera hoja con comparación detallada modelo vs golden standard."""
        comparaciones = []
        for resultado in self.resultados:
            for comp in resultado.get("comparaciones_evaluadas", []):
                if "referencia_ponderada" in comp:
                    comparaciones.append({
                        "Archivo": resultado['nombre_archivo'],
                        "ID Origen": comp['idCursoOrigen'],
                        "ID Destino": comp['idCursoDestino'],
                        "Modelo Similitud": comp['similitud_ponderada'],
                        "Golden Similitud": comp['referencia_ponderada'],
                        "Diferencia Absoluta": comp['diferencia'],
                        "Modelo Decisión": "Similar" if self.es_similar(comp['similitud_ponderada']) else "No Similar",
                        "Golden Decisión": "Similar" if self.es_similar(comp['referencia_ponderada']) else "No Similar",
                        "Concordancia": "Sí" if comp['acierto'] else "No",
                        "Tipo Error": "Falso Positivo" if (self.es_similar(comp['similitud_ponderada']) and
                                                          not self.es_similar(comp['referencia_ponderada']))
                                      else "Falso Negativo" if (not self.es_similar(comp['similitud_ponderada']) and
                                                              self.es_similar(comp['referencia_ponderada']))
                                      else "N/A"
                    })

        if not comparaciones:
            df = pd.DataFrame({"Mensaje": ["No hay datos de comparación con golden standard"]})
            self.report_generator.add_sheet(
                title="Comparación Modelo vs Golden",
                df=df
            )
            return

        df = pd.DataFrame(comparaciones)

        # Generar gráfico de dispersión
        plt.figure(figsize=(10, 10))
        sns.scatterplot(
            x='Golden Similitud',
            y='Modelo Similitud',
            hue='Concordancia',
            data=df,
            palette={'Sí': 'green', 'No': 'red'}
        )

        plt.plot([0, 1], [0, 1], 'r--')
        plt.axhline(y=THRESHOLD_SIMILITUD_PONDERADA, color='gray', linestyle='--')
        plt.axvline(x=THRESHOLD_SIMILITUD_PONDERADA, color='gray', linestyle='--')
        plt.title("Modelo vs Golden Standard")
        plt.xlabel("Golden Standard Similitud")
        plt.ylabel("Modelo Similitud")

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png')
        img_buffer.seek(0)
        plt.close()

        # Añadir hoja con datos y gráfico
        self.report_generator.add_sheet(
            title="Comparación Modelo vs Golden",
            df=df,
            notes=[
                "Comparación detallada entre predicciones del modelo y golden standard",
                f"Umbral de similitud: {THRESHOLD_SIMILITUD_PONDERADA}",
                f"Total comparaciones: {len(comparaciones)}",
                f"Tasa de concordancia: {self.metricas_globales.get('tasa_acierto', 0) * 100:.2f}%"
            ],
            conditional_cols={
                "Diferencia Absoluta": [
                    {'type': 'less', 'value': 0.1, 'fill': 'green'},
                    {'type': 'range', 'min': 0.1, 'max': 0.2, 'fill': 'yellow'},
                    {'type': 'greater_equal', 'value': 0.2, 'fill': 'red'}
                ],
                "Concordancia": [
                    {'type': 'equal', 'value': 'Sí', 'fill': 'green'},
                    {'type': 'equal', 'value': 'No', 'fill': 'red'}
                ]
            }
        )

        self.report_generator.add_image(
            "Comparación Modelo vs Golden",
            img_buffer.getvalue(),
            'J2'
        )

    def _generar_matriz_confusion(self):
        """Genera hoja con matriz de confusión y reporte de clasificación."""
        y_true = []
        y_pred = []
        for resultado in self.resultados:
            for comp in resultado.get("comparaciones_evaluadas", []):
                if "referencia_ponderada" in comp:
                    y_true.append(1 if self.es_similar(comp['referencia_ponderada']) else 0)
                    y_pred.append(1 if self.es_similar(comp['similitud_ponderada']) else 0)

        if not y_true:
            df = pd.DataFrame({"Mensaje": ["No hay datos suficientes para generar matriz de confusión"]})
            self.report_generator.add_sheet(
                title="Matriz de Confusión",
                df=df
            )
            return

        # Get all possible labels (even if not present in data)
        labels = [0, 1]

        # Matriz de confusión
        cm = confusion_matrix(y_true, y_pred, labels=labels)

        # Handle case where all predictions are the same
        if len(np.unique(y_true)) == 1 and len(np.unique(y_pred)) == 1:
            # If all are 0
            if y_true[0] == 0:
                cm = np.array([[len(y_true), 0], [0, 0]])
            # If all are 1
            else:
                cm = np.array([[0, 0], [0, len(y_true)]])

        cm_df = pd.DataFrame(
            cm,
            index=["Real No Similar", "Real Similar"],
            columns=["Pred No Similar", "Pred Similar"]
        )

        # Reporte de clasificación
        report = classification_report(y_true, y_pred, output_dict=True, zero_division=0)
        report_df = pd.DataFrame(report).transpose()

        # Visualización matriz de confusión
        plt.figure(figsize=(8, 6))
        sns.heatmap(cm_df, annot=True, fmt='d', cmap='Blues')
        plt.title("Matriz de Confusión")
        plt.ylabel("Real")
        plt.xlabel("Predicho")

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png')
        img_buffer.seek(0)
        plt.close()

        # Añadir hojas al reporte
        self.report_generator.add_sheet(
            title="Matriz de Confusión",
            df=cm_df,
            notes=[
                "Comparación entre decisiones del modelo y golden standard",
                f"Umbral de similitud: {THRESHOLD_SIMILITUD_PONDERADA}"
            ]
        )

        self.report_generator.add_image(
            "Matriz de Confusión",
            img_buffer.getvalue(),
            'D10'
        )

        self.report_generator.add_sheet(
            title="Reporte Clasificación",
            df=report_df,
            notes=[
                "Métricas de clasificación del modelo vs golden standard",
                "Precision: Verdaderos positivos / (Verdaderos positivos + Falsos positivos)",
                "Recall: Verdaderos positivos / (Verdaderos positivos + Falsos negativos)",
                "F1-score: Media armónica de precision y recall"
            ]
        )

    def _generar_analisis_errores_detallado(self):
        """Genera hoja con análisis detallado de errores."""
        errores = []
        for resultado in self.resultados:
            for comp in resultado.get("comparaciones_evaluadas", []):
                if "referencia_ponderada" in comp and not comp['acierto']:
                    # Calcular diferencias por componente
                    dif_sumilla = abs(comp.get('similitud_sumilla', 0) -
                                      self.golden_standard.get(
                                          f"{comp['idCursoOrigen']}_{comp['idCursoDestino']}", {}
                                      ).get('similitud_sumilla', 0))
                    dif_aprendizajes = abs(comp.get('similitud_aprendizajes', 0) -
                                           self.golden_standard.get(
                                               f"{comp['idCursoOrigen']}_{comp['idCursoDestino']}", {}
                                           ).get('similitud_aprendizajes', 0))
                    dif_unidades = abs(comp.get('similitud_unidades', 0) -
                                       self.golden_standard.get(
                                           f"{comp['idCursoOrigen']}_{comp['idCursoDestino']}", {}
                                       ).get('similitud_unidades', 0))
                    dif_bibliografia = abs(comp.get('similitud_bibliografia', 0) -
                                           self.golden_standard.get(
                                               f"{comp['idCursoOrigen']}_{comp['idCursoDestino']}", {}
                                           ).get('similitud_bibliografia', 0))

                    errores.append({
                        "Archivo": resultado['nombre_archivo'],
                        "ID Origen": comp['idCursoOrigen'],
                        "ID Destino": comp['idCursoDestino'],
                        "Tipo Error": "Falso Positivo" if comp[
                                                              'similitud_ponderada'] >= THRESHOLD_SIMILITUD_PONDERADA else "Falso Negativo",
                        "Modelo Similitud": comp['similitud_ponderada'],
                        "Golden Similitud": comp['referencia_ponderada'],
                        "Diferencia Total": comp['diferencia'],
                        "Dif. Sumilla": dif_sumilla,
                        "Dif. Aprendizajes": dif_aprendizajes,
                        "Dif. Unidades": dif_unidades,
                        "Dif. Bibliografía": dif_bibliografia,
                        "Componente Mayor Error": max(
                            [("Sumilla", dif_sumilla),
                             ("Aprendizajes", dif_aprendizajes),
                             ("Unidades", dif_unidades),
                             ("Bibliografía", dif_bibliografia)],
                            key=lambda x: abs(x[1]))[0]
                    })

        if not errores:
            df = pd.DataFrame({"Mensaje": ["No se encontraron errores en las comparaciones"]})
            self.report_generator.add_sheet(
                title="Análisis de Errores",
                df=df
            )
            return

        df = pd.DataFrame(errores)

        # Análisis agregado de errores
        analisis_errores = {
            "Tipo Error": ["Falso Positivo", "Falso Negativo", "Total"],
            "Cantidad": [
                len([e for e in errores if e['Tipo Error'] == "Falso Positivo"]),
                len([e for e in errores if e['Tipo Error'] == "Falso Negativo"]),
                len(errores)
            ],
            "Porcentaje": [
                len([e for e in errores if e['Tipo Error'] == "Falso Positivo"]) / len(errores) * 100,
                len([e for e in errores if e['Tipo Error'] == "Falso Negativo"]) / len(errores) * 100,
                100
            ]
        }
        df_analisis = pd.DataFrame(analisis_errores)

        # Componentes con mayor error
        componentes_error = pd.DataFrame(
            df['Componente Mayor Error'].value_counts()
        ).reset_index()
        componentes_error.columns = ['Componente', 'Errores']
        componentes_error['Porcentaje'] = componentes_error['Errores'] / len(errores) * 100

        # Añadir hojas al reporte
        self.report_generator.add_sheet(
            title="Análisis de Errores",
            df=df,
            notes=[
                "Análisis detallado de los casos con discrepancia",
                f"Total errores: {len(errores)}",
                f"Umbral similitud: {THRESHOLD_SIMILITUD_PONDERADA}"
            ],
            conditional_cols={
                "Diferencia Total": [
                    {'type': 'greater_equal', 'value': 0.3, 'fill': 'red'},
                    {'type': 'range', 'min': 0.2, 'max': 0.3, 'fill': 'yellow'},
                    {'type': 'less', 'value': 0.2, 'fill': 'green'}
                ]
            }
        )

        self.report_generator.add_sheet(
            title="Resumen Errores",
            df=df_analisis,
            chart_data={
                'type': 'bar',
                'title': 'Distribución de Tipos de Error',
                'y_title': 'Porcentaje',
                'pos': 'E2'
            }
        )

        self.report_generator.add_sheet(
            title="Componentes con Error",
            df=componentes_error,
            chart_data={
                'type': 'bar',
                'title': 'Componentes con Mayor Error',
                'y_title': 'Porcentaje',
                'pos': 'E2'
            }
        )

    def _generar_detalle_por_archivo(self):
        """Genera hoja con detalle de métricas por archivo."""
        data = []
        for resultado in self.resultados:
            if not resultado:
                continue

            data.append({
                "Archivo": resultado.get('nombre_archivo', ''),
                "Comparaciones": resultado.get('num_comparaciones', 0),
                "Tiempo (s)": resultado.get('tiempo_respuesta', 0),
                "Similitud Promedio": resultado.get('metricas_similitud', {}).get('ponderada', {}).get('mean', 0),
                "Emparejamiento Promedio": resultado.get('metricas_emparejamiento', {}).get(
                    'mean_porcentaje_emparejamiento', 0),
                "Coherencia Interna": resultado.get('metricas_coherencia', {}).get('coherencia_interna', 0)
            })

        if not data:
            df = pd.DataFrame({"Mensaje": ["No hay datos disponibles para generar esta hoja"]})
        else:
            df = pd.DataFrame(data)

        self.report_generator.add_sheet(
            title="Detalle por Archivo",
            df=df,
            chart_data={
                'type': 'line',
                'title': 'Evolución Similitud por Archivo',
                'y_title': 'Similitud',
                'pos': 'A15'
            } if len(df) > 1 else None
        )

    def _generar_comparaciones_individuales(self):
        """Genera hoja con todas las comparaciones individuales."""
        data = []
        for resultado in self.resultados:
            if not resultado:
                continue

            for comp in resultado.get("comparaciones_evaluadas", []):
                row = {
                    "Archivo": resultado.get('nombre_archivo', ''),
                    "ID Origen": comp.get('idCursoOrigen', ''),
                    "ID Destino": comp.get('idCursoDestino', ''),
                    "Similitud Sumilla": comp.get('similitud_sumilla', 0),
                    "Similitud Aprendizajes": comp.get('similitud_aprendizajes', 0),
                    "Similitud Unidades": comp.get('similitud_unidades', 0),
                    "Similitud Bibliografía": comp.get('similitud_bibliografia', 0),
                    "Similitud Ponderada": comp.get('similitud_ponderada', 0),
                    "Es Similar": "Sí" if comp.get('similitud_ponderada', 0) >= THRESHOLD_SIMILITUD_PONDERADA else "No"
                }

                if 'referencia_ponderada' in comp:
                    row.update({
                        "Referencia Ponderada": comp['referencia_ponderada'],
                        "Diferencia": comp['diferencia'],
                        "Acierto": "Sí" if comp['acierto'] else "No"
                    })

                data.append(row)

        if not data:
            df = pd.DataFrame({"Mensaje": ["No hay comparaciones individuales para mostrar"]})
        else:
            df = pd.DataFrame(data)

        self.report_generator.add_sheet(
            title="Comparaciones Individuales",
            df=df,
            notes=["Detalle de todas las comparaciones realizadas"]
        )


def main():
    """Función principal para ejecutar el evaluador."""
    try:
        evaluador = SemanticPrecisionEvaluator()
        evaluador.ejecutar_evaluacion()
    except Exception as e:
        logger.error(f"Error en la ejecución principal: {str(e)}")
        raise


if __name__ == '__main__':
    main()